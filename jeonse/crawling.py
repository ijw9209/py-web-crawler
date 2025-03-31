import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os  # 🔹 폴더 생성용 모듈 추가

# 현재 날짜 및 시간
now = datetime.now().strftime("%Y%m%d_%H%M")

# 🔹 저장할 폴더 설정
output_folder = "output"
os.makedirs(output_folder, exist_ok=True)  # 폴더 없으면 생성

# 지역 코드 목록
regions = {"서울": "01", "경기": "07"}

# 크롤링할 기본 URL
base_url = "https://www.khug.or.kr/jeonse/web/s07/s070102.jsp"
detail_base_url = "https://www.khug.or.kr/jeonse/web/s07/"

# 지역별 크롤링 실행
for region_name, sido_code in regions.items():
    all_data = []
    first_page_url = f"{base_url}?sbGugun=ALL&CMB_SIDO={sido_code}&cur_page=1"
    response = requests.get(first_page_url)

    if response.status_code == 200:
        soup = BeautifulSoup(response.text, 'html.parser')
        last_page_btn = soup.find("a", class_="last")
        last_page = int(last_page_btn["href"].split("cur_page=")[-1]) if last_page_btn else 1
        print(f"{region_name} 지역 크롤링 시작 (총 {last_page} 페이지)")

        for page in range(1, last_page + 1):
            page_url = f"{base_url}?sbGugun=ALL&CMB_SIDO={sido_code}&cur_page={page}"
            response = requests.get(page_url)

            if response.status_code == 200:
                soup = BeautifulSoup(response.text, 'html.parser')
                table = soup.find('table')

                if table:
                    rows = table.find_all('tr')
                    for row in rows:
                        cols = row.find_all('td')
                        if cols:
                            address_link = cols[5].find("a")
                            address_href = detail_base_url + address_link["href"] if address_link else ""
                            address_text = address_link.text.strip() if address_link else cols[5].text.strip()
                            cols_text = [col.text.strip() for col in cols]
                            cols_text[5] = (address_text, address_href)  # 주소 하이퍼링크 추가
                            all_data.append(cols_text)

    df = pd.DataFrame(all_data, columns=["번호", "공고일자", "청약 접수기간", "시도", "시군구", "주소", "주택유형", "전용면적(m2)", "임대보증금액", "신청자수"])
    # 🔹 파일 경로 설정 (output 폴더에 저장)
    file_name = f"크롤링_데이터_{region_name}_{now}.xlsx"
    file_path = os.path.join(output_folder, file_name)

    df.to_excel(file_path, index=False, engine='openpyxl')

    # 🔹 엑셀 파일 열기
    wb = load_workbook(file_path)
    ws = wb.active

    # 🔹 주소 컬럼 하이퍼링크 적용
    for row_idx, (text, link) in enumerate(df["주소"], start=2):
        cell = ws[f"{get_column_letter(6)}{row_idx}"]
        cell.value = text
        if link:
            cell.hyperlink = link
            cell.style = "Hyperlink"

    # 🔹 특정 컬럼(주소, 시군구)만 너비 조정
    ws.column_dimensions[get_column_letter(5)].width = 20  # 시군구
    ws.column_dimensions[get_column_letter(6)].width = 60  # 주소

    wb.save(file_path)
    print(f"{region_name} 지역 크롤링 완료! 엑셀 저장: {file_path}")

print("모든 지역 크롤링 완료!")
